from django.shortcuts import render
from django.http import HttpResponse
from .forms import UploadFileForm
import pandas as pd
from openpyxl import load_workbook

def handle_uploaded_file(f):
    path = f'temp/{f.name}'
    with open(path, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    return path

def fill_out_report(sad_path, budget_path):
    # Load the report template
    report_template_path = 'acc_report/static/acc_report.xlsx'
    wb = load_workbook(report_template_path)
    ws = wb.active

    # Load the Excel files
    sad_df = pd.read_excel(sad_path)
    budget_df = pd.read_excel(budget_path)

    # Example: Fill out the report using the data
    ws['A1'] = sad_df.iloc[0, 0]
    ws['B1'] = budget_df.iloc[0, 0]
    # Add more logic to fill out the rest of the report

    # Save the filled report
    report_path = 'temp/Acc_Report.xlsx'
    wb.save(report_path)
    return report_path

def upload_files(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            sad_path = handle_uploaded_file(request.FILES['sad_file'])
            budget_path = handle_uploaded_file(request.FILES['budget_file'])

            report_path = fill_out_report(sad_path, budget_path)

            # Serve the generated report file
            with open(report_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                response['Content-Disposition'] = 'inline; filename=' + 'Acc_Report.xlsx'
                return response
    else:
        form = UploadFileForm()

    return render(request, 'acc_upload.html', {'form': form})